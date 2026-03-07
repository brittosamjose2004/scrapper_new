import argparse
import json
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def flatten_esg_json(data):
    metadata = data.get("metadata", {})
    results = data.get("esg_results", {})

    rows = []
    for pillar_key, pillar_data in results.items():
        pillar_label = pillar_key.replace("_", " ").title()

        for subcategory_key, subcategory_data in pillar_data.items():
            subcategory_label = subcategory_data.get("label", subcategory_key.replace("_", " ").title())
            metrics = subcategory_data.get("metrics", [])

            for item in metrics:
                answer = item.get("answer", {})
                rows.append({
                    "company": metadata.get("company"),
                    "analysis_date": metadata.get("analysis_date"),
                    "pillar_key": pillar_key,
                    "pillar": pillar_label,
                    "subcategory_key": subcategory_key,
                    "subcategory": subcategory_label,
                    "question": item.get("question"),
                    "value": answer.get("value"),
                    "unit": answer.get("unit"),
                    "year": answer.get("year"),
                    "previous_year_value": answer.get("previous_year_value"),
                    "source_detail": answer.get("source_detail"),
                    "confidence": answer.get("confidence"),
                    "sources_searched": item.get("sources_searched"),
                })

    return metadata, rows


def autosize_columns(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)


def write_metadata_sheet(wb, metadata):
    ws = wb.active
    ws.title = "Metadata"

    ws.append(["Field", "Value"])
    for key, value in metadata.items():
        if isinstance(value, dict):
            ws.append([key, json.dumps(value, ensure_ascii=False)])
        else:
            ws.append([key, value])

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    autosize_columns(ws)


def write_metrics_sheet(wb, rows):
    ws = wb.create_sheet("Metrics")

    columns = [
        "company",
        "analysis_date",
        "pillar_key",
        "pillar",
        "subcategory_key",
        "subcategory",
        "question",
        "value",
        "unit",
        "year",
        "previous_year_value",
        "confidence",
        "sources_searched",
        "source_detail",
    ]

    ws.append(columns)
    for row in rows:
        ws.append([row.get(col) for col in columns])

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.freeze_panes = "A2"
    autosize_columns(ws)


def convert_json_to_excel(input_json, output_xlsx=None):
    with open(input_json, "r", encoding="utf-8") as f:
        data = json.load(f)

    metadata, rows = flatten_esg_json(data)

    if output_xlsx is None:
        base, _ = os.path.splitext(input_json)
        output_xlsx = f"{base}.xlsx"

    wb = Workbook()
    write_metadata_sheet(wb, metadata)
    write_metrics_sheet(wb, rows)
    wb.save(output_xlsx)

    return output_xlsx, len(rows)


def main():
    parser = argparse.ArgumentParser(description="Convert ESG JSON output to structured Excel file")
    parser.add_argument("--input", required=True, help="Path to ESG answers JSON file")
    parser.add_argument("--output", default=None, help="Optional output .xlsx path")
    args = parser.parse_args()

    output_path, row_count = convert_json_to_excel(args.input, args.output)
    print(f"✅ Excel created: {output_path}")
    print(f"📊 Metrics rows: {row_count}")
    print(f"🕒 Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()

import argparse
import json
import os
import re
import time
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from esg_analyzer import (
    ESG_QUESTIONS_FILE,
    ESGAnalyzer,
    MultiProviderLLM,
    GeminiRestLLM,
    OpenAICompatLLM,
    OllamaLLM,
    GROK_MODEL,
    OPENROUTER_MODEL,
    _split_keys,
    _sanitize,
    build_prompt,
    build_context,
    find_relevant_chunks,
    load_all_data,
)


def extract_years(text):
    years = set()
    if not text:
        return years

    for match in re.findall(r"\b(19\d{2}|20\d{2})\b", text):
        y = int(match)
        if 2000 <= y <= 2035:
            years.add(y)

    for match in re.findall(r"\bFY\s*'?([0-9]{2})\b", text, flags=re.IGNORECASE):
        yy = int(match)
        years.add(2000 + yy)

    return years


def infer_chunk_years(chunk):
    source = chunk.get("source", "")
    text = chunk.get("text", "")[:1200]
    years = extract_years(source)
    if not years:
        years = extract_years(text)
    return sorted(years)


def group_chunks_by_year(all_chunks):
    grouped = {}
    for chunk in all_chunks:
        years = infer_chunk_years(chunk)
        if not years:
            continue
        for year in years:
            grouped.setdefault(year, []).append(chunk)
    return grouped


def group_annual_report_chunks_by_year(all_chunks):
    grouped = {}
    for chunk in all_chunks:
        source = chunk.get("source", "")
        source_file = source.split(",", 1)[0].strip().lower()
        if "annual report" not in source_file and "annualreport" not in source_file:
            continue

        year = None
        m = re.match(r"\s*(20\d{2})", source_file)
        if m:
            year = int(m.group(1))
        else:
            ys = extract_years(source_file)
            if ys:
                year = sorted(ys)[0]

        if year:
            grouped.setdefault(year, []).append(chunk)
    return grouped


def get_annual_report_years(base_folder, company_name):
    sanitized = _sanitize(company_name)
    candidate_dirs = [
        os.path.join(base_folder, "nseindia.com", sanitized),
        os.path.join(base_folder, "annualreports.com", sanitized),
        os.path.join(base_folder, sanitized),
    ]

    years = set()
    for folder in candidate_dirs:
        if not os.path.exists(folder):
            continue
        for root, _, files in os.walk(folder):
            for fname in files:
                low = fname.lower()
                if not low.endswith(".pdf"):
                    continue
                if "annual report" not in low and "annualreport" not in low:
                    continue
                match = re.match(r"\s*(20\d{2})", fname)
                if match:
                    years.add(int(match.group(1)))

    return sorted(years)


def build_llm(llm_type="auto", llm_chain=None, gemini_key=None, grok_key=None, openrouter_key=None):
    configured_chain = llm_chain or os.environ.get("LLM_CHAIN", "")
    provider_order = [p.strip().lower() for p in configured_chain.split(",") if p.strip()] if configured_chain else [llm_type.lower()]
    if provider_order == ["auto"]:
        provider_order = ["gemini", "grok", "openrouter", "ollama"]

    gemini_keys = _split_keys(gemini_key) or _split_keys(os.environ.get("GEMINI_API_KEY", ""))
    grok_keys = _split_keys(grok_key) or _split_keys(os.environ.get("XAI_API_KEY", ""))
    openrouter_keys = _split_keys(openrouter_key) or _split_keys(os.environ.get("OPENROUTER_API_KEY", ""))

    clients = []
    for provider in provider_order:
        if provider == "gemini":
            for idx, key in enumerate(gemini_keys, start=1):
                clients.append((f"gemini#{idx}", GeminiRestLLM(api_key=key)))
        elif provider == "grok":
            for idx, key in enumerate(grok_keys, start=1):
                clients.append((
                    f"grok#{idx}",
                    OpenAICompatLLM(
                        api_key=key,
                        base_url="https://api.x.ai/v1",
                        model=GROK_MODEL,
                        provider_name="Grok",
                    ),
                ))
        elif provider == "openrouter":
            for idx, key in enumerate(openrouter_keys, start=1):
                clients.append((
                    f"openrouter#{idx}",
                    OpenAICompatLLM(
                        api_key=key,
                        base_url="https://openrouter.ai/api/v1",
                        model=OPENROUTER_MODEL,
                        provider_name="OpenRouter",
                    ),
                ))
        elif provider == "ollama":
            try:
                clients.append(("ollama", OllamaLLM()))
            except SystemExit:
                pass

    if not clients:
        raise RuntimeError("No valid LLM provider configured.")

    return MultiProviderLLM(clients) if len(clients) > 1 else clients[0][1], [c[0] for c in clients]


def search_online_context(company_name, question, year, max_results=4):
    query = f"{company_name} {question} {year} annual report sustainability"
    url = "https://html.duckduckgo.com/html/"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    }
    snippets = []

    try:
        resp = requests.post(url, data={"q": query}, headers=headers, timeout=8)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        links = soup.select("a.result__a")

        for link in links[:max_results]:
            title = link.get_text(" ", strip=True)
            href = link.get("href", "")
            container = link.find_parent("div", class_="result")
            snippet_node = container.select_one("a.result__snippet") if container else None
            snippet = snippet_node.get_text(" ", strip=True) if snippet_node else ""
            snippets.append(f"[Online: {title}]\nURL: {href}\nSnippet: {snippet}")
    except BaseException:
        # Catch everything including KeyboardInterrupt and network failures
        return ""

    return "\n\n---\n\n".join(snippets)


def is_missing_answer(parsed):
    value = str(parsed.get("value", "")).strip().lower()
    conf = str(parsed.get("confidence", "low")).strip().lower()
    return value in {"", "not disclosed", "none", "null", "n/a", "na", "[error]"} or conf == "low"


def run_yearly_analysis(
    company_name,
    base_folder="downloads",
    questions_path=None,
    llm_type="auto",
    llm_chain="gemini,grok,openrouter",
    gemini_key=None,
    grok_key=None,
    openrouter_key=None,
    years=None,
    max_years=3,
    online_fallback=True,
    sleep_seconds=1.0,
    annual_only=True,
):
    print("\n" + "=" * 80)
    print("📅 YEAR-WISE ESG ANALYSIS")
    print("=" * 80)

    print(f"\n  📂 Loading all scraped data for '{company_name}'...")
    all_chunks, summary = load_all_data(base_folder, company_name)
    if not all_chunks:
        raise RuntimeError("No chunks found from scraped data.")

    grouped = group_chunks_by_year(all_chunks)
    annual_grouped = group_annual_report_chunks_by_year(all_chunks)
    annual_report_years = get_annual_report_years(base_folder, company_name)
    if not annual_report_years:
        discovered_years = sorted(grouped.keys())
        if not discovered_years:
            raise RuntimeError("Could not infer report years from sources.")
        annual_report_years = discovered_years

    if years:
        target_years = sorted([int(y) for y in years if int(y) in annual_report_years])
    else:
        if max_years and max_years > 0:
            target_years = annual_report_years[-max_years:]
        else:
            target_years = annual_report_years

    if not target_years:
        raise RuntimeError("No matching target years with available chunks.")

    if questions_path is None:
        questions_path = ESG_QUESTIONS_FILE
    with open(questions_path, "r", encoding="utf-8") as f:
        questions_data = json.load(f)

    llm, providers = build_llm(
        llm_type=llm_type,
        llm_chain=llm_chain,
        gemini_key=gemini_key,
        grok_key=grok_key,
        openrouter_key=openrouter_key,
    )
    print(f"\n  ✅ LLM providers ready: {', '.join(providers)}")
    print(f"  📆 Target years: {', '.join(str(y) for y in target_years)}")

    sanitized = _sanitize(company_name)
    out_folder = os.path.join(base_folder, "nseindia.com", sanitized)
    if not os.path.exists(out_folder):
        out_folder = os.path.join(base_folder, sanitized)
    os.makedirs(out_folder, exist_ok=True)
    date_str = datetime.now().strftime("%Y%m%d")

    all_year_results = {}
    per_year_files = {}
    start_all = time.time()
    total_questions = sum(len(sub.get("questions", [])) for pillar in questions_data.values() for sub in pillar.values())

    # Pre-compute BRSR / sustainability-report chunks for augmentation
    _brsr_kws = ("brsr", "sustainability", "tcfd", "csr")
    brsr_chunks = [c for c in all_chunks
                   if any(k in c.get("source", "").lower() for k in _brsr_kws)]

    for year in target_years:
        print("\n" + "-" * 80)
        year_chunks = annual_grouped.get(year, []) if annual_only else grouped.get(year, [])
        chunk_label = "annual-report chunks" if annual_only else "year-specific chunks"

        # Augment sparse years: if fewer than 80 chunks, also add BRSR/sustainability docs
        # (these are often multi-year reference docs covering recent data)
        if len(year_chunks) < 80 and brsr_chunks:
            existing_ids = {id(c) for c in year_chunks}
            added = [c for c in brsr_chunks if id(c) not in existing_ids]
            year_chunks = year_chunks + added
            chunk_label += f" (+{len(added)} BRSR/sustainability)"

        # Further fallback: if still very sparse, use ALL scraped chunks
        if len(year_chunks) < 30 and all_chunks:
            year_chunks = all_chunks
            chunk_label = "all scraped chunks (fallback)"

        print(f"🗓️  Processing Year {year} ({len(year_chunks)} {chunk_label})")
        print("-" * 80)

        analyzer = ESGAnalyzer(llm, year_chunks, company_name, summary)
        analyzer.total_questions = total_questions
        analyzer.answered = 0
        year_results = {}

        TOP_K = int(os.environ.get("TOP_K_CHUNKS", "10"))
        MAX_CTX = int(os.environ.get("MAX_CTX_CHARS", "14000"))

        for pillar_key, pillar_data in questions_data.items():
            pillar_label = pillar_key.replace("_", " ").title()
            print(f"\n  📌 {pillar_label}")
            year_results[pillar_key] = {}

            for sub_key, sub_data in pillar_data.items():
                sub_label = sub_key.replace("_", " ").title()
                questions = sub_data.get("questions", [])
                print(f"    📂 {sub_label} ({len(questions)} metrics)")
                metrics = []

                for q in questions:
                    analyzer.answered += 1
                    progress = f"[{analyzer.answered}/{analyzer.total_questions}]"
                    display_q = q[:55] + "..." if len(q) > 55 else q
                    print(f"      {progress} {display_q}", end=" ", flush=True)

                    relevant = find_relevant_chunks(q, year_chunks, top_k=TOP_K)
                    local_context = build_context(relevant, max_chars=MAX_CTX)

                    # If context is still empty, search all available chunks
                    if not local_context.strip() and year_chunks is not all_chunks:
                        fallback_rel = find_relevant_chunks(q, all_chunks, top_k=TOP_K)
                        local_context = build_context(fallback_rel, max_chars=MAX_CTX)
                        relevant = fallback_rel

                    parsed = {
                        "value": "Not disclosed",
                        "unit": "",
                        "year": year,
                        "previous_year_value": None,
                        "source_detail": "No relevant data found in year-specific sources; checked all scraped sources",
                        "confidence": "low",
                    }

                    if local_context.strip():
                        prompt = build_prompt(q, pillar_label, sub_label, local_context, company_name)
                        raw = llm.generate(prompt)
                        parsed = analyzer._parse_llm_answer(raw)

                    used_online = False
                    if online_fallback and is_missing_answer(parsed):
                        try:
                            online_context = search_online_context(company_name, q, year)
                            if online_context.strip():
                                online_prompt = build_prompt(q, pillar_label, sub_label, online_context, company_name)
                                online_raw = llm.generate(online_prompt)
                                online_parsed = analyzer._parse_llm_answer(online_raw)
                                if not is_missing_answer(online_parsed):
                                    parsed = online_parsed
                                    used_online = True
                                    src_detail = str(parsed.get("source_detail", "")).strip()
                                    parsed["source_detail"] = (src_detail + " | Filled via online search context").strip(" |")
                        except Exception:
                            pass  # Network not available — skip silently

                    if not parsed.get("year"):
                        parsed["year"] = year

                    conf = parsed.get("confidence", "low")
                    val = str(parsed.get("value", "?"))[:30]
                    suffix = " +online" if used_online else ""
                    print(f"→ {val} ({conf}){suffix}")

                    metrics.append({
                        "question": q,
                        "answer": parsed,
                        "sources_searched": len(relevant),
                        "online_fallback_used": used_online,
                    })

                    time.sleep(max(0.0, sleep_seconds))

                year_results[pillar_key][sub_key] = {
                    "label": sub_label,
                    "metrics": metrics,
                }

        all_year_results[str(year)] = year_results

        year_payload = {
            "metadata": {
                "company": company_name,
                "analysis_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "mode": "yearly-single",
                "year": year,
                "llm_providers": providers,
                "online_fallback": online_fallback,
                "annual_only_mode": annual_only,
            },
            "esg_results": year_results,
        }
        year_json_path = os.path.join(out_folder, f"{sanitized}_ESG_Answers_{year}_{date_str}.json")
        with open(year_json_path, "w", encoding="utf-8") as yf:
            json.dump(year_payload, yf, indent=2, ensure_ascii=False)
        per_year_files[str(year)] = year_json_path
        print(f"  💾 Saved year file: {year_json_path}")

    elapsed = time.time() - start_all
    out_json = os.path.join(out_folder, f"{sanitized}_ESG_Answers_Yearly_{date_str}.json")

    output = {
        "metadata": {
            "company": company_name,
            "analysis_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "mode": "yearly",
            "years_processed": target_years,
            "total_years": len(target_years),
            "time_taken_seconds": round(elapsed, 1),
            "llm_providers": providers,
            "online_fallback": online_fallback,
            "annual_only_mode": annual_only,
            "per_year_json_files": {},
            "data_sources": {
                "pdf_files": summary.get("pdf_count", 0),
                "pdf_pages_extracted": summary.get("pdf_pages", 0),
                "news_articles": summary.get("news_items", 0),
                "social_media_posts": summary.get("social_items", 0),
                "total_text_chunks": len(all_chunks),
            },
        },
        "yearly_esg_results": all_year_results,
    }

    output["metadata"]["per_year_json_files"] = per_year_files

    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    out_xlsx = out_json.replace(".json", ".xlsx")
    write_yearly_excel(output, out_xlsx)

    return out_json, out_xlsx, per_year_files


def write_yearly_excel(output, out_xlsx):
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "Metadata"
    ws_meta.append(["Field", "Value"])

    for k, v in output.get("metadata", {}).items():
        if isinstance(v, (dict, list)):
            ws_meta.append([k, json.dumps(v, ensure_ascii=False)])
        else:
            ws_meta.append([k, v])

    ws = wb.create_sheet("Yearly_Metrics")
    columns = [
        "year",
        "pillar",
        "subcategory",
        "question",
        "value",
        "unit",
        "answer_year",
        "previous_year_value",
        "confidence",
        "sources_searched",
        "online_fallback_used",
        "source_detail",
    ]
    ws.append(columns)

    yearly = output.get("yearly_esg_results", {})
    for year, pillar_data in yearly.items():
        for pillar_key, sub_data in pillar_data.items():
            pillar_label = pillar_key.replace("_", " ").title()
            for sub_key, sub_block in sub_data.items():
                sub_label = sub_block.get("label", sub_key.replace("_", " ").title())
                for item in sub_block.get("metrics", []):
                    ans = item.get("answer", {})
                    ws.append([
                        year,
                        pillar_label,
                        sub_label,
                        item.get("question"),
                        ans.get("value"),
                        ans.get("unit"),
                        ans.get("year"),
                        ans.get("previous_year_value"),
                        ans.get("confidence"),
                        item.get("sources_searched"),
                        item.get("online_fallback_used", False),
                        ans.get("source_detail"),
                    ])

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in [ws_meta, ws]:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"

    for sheet in [ws_meta, ws]:
        for col_cells in sheet.columns:
            max_len = 0
            letter = get_column_letter(col_cells[0].column)
            for c in col_cells:
                txt = "" if c.value is None else str(c.value)
                max_len = max(max_len, len(txt))
            sheet.column_dimensions[letter].width = min(max_len + 2, 65)

    wb.save(out_xlsx)


def main():
    parser = argparse.ArgumentParser(description="Year-wise ESG question answering with online fallback")
    parser.add_argument("--company", required=True)
    parser.add_argument("--folder", default="downloads")
    parser.add_argument("--questions", default=None)

    parser.add_argument("--llm", default="auto", help="gemini|grok|openrouter|ollama|auto")
    parser.add_argument("--llm-chain", default="gemini,grok,openrouter")
    parser.add_argument("--gemini-key", default=None)
    parser.add_argument("--grok-key", default=None)
    parser.add_argument("--openrouter-key", default=None)

    parser.add_argument("--years", default=None, help="Comma-separated years, e.g. 2023,2024,2025")
    parser.add_argument("--max-years", type=int, default=0, help="Used when --years is not provided; 0 means all annual-report years")
    parser.add_argument("--no-online-fallback", action="store_true", help="Disable web fallback for missing answers")
    parser.add_argument("--include-all-scraped", action="store_true", help="Use all scraped year-matched chunks, not only annual-report chunks")
    parser.add_argument("--sleep-seconds", type=float, default=1.0)

    args = parser.parse_args()

    years = None
    if args.years:
        years = [y.strip() for y in args.years.split(",") if y.strip()]

    out_json, out_xlsx, per_year_files = run_yearly_analysis(
        company_name=args.company,
        base_folder=args.folder,
        questions_path=args.questions,
        llm_type=args.llm,
        llm_chain=args.llm_chain,
        gemini_key=args.gemini_key,
        grok_key=args.grok_key,
        openrouter_key=args.openrouter_key,
        years=years,
        max_years=args.max_years,
        online_fallback=not args.no_online_fallback,
        sleep_seconds=args.sleep_seconds,
        annual_only=not args.include_all_scraped,
    )

    print("\n" + "=" * 80)
    print("✅ YEAR-WISE ESG ANALYSIS COMPLETE")
    print(f"📄 JSON: {out_json}")
    print(f"📊 Excel: {out_xlsx}")
    print(f"🧩 Per-year JSON files: {len(per_year_files)}")
    print("=" * 80)


if __name__ == "__main__":
    main()
